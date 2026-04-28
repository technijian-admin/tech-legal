#!/usr/bin/env python3
"""Build master BWH life-of-contract xlsx for Dave Barisic.

Structure:
  Sheet 1: Summary (invoice footer reconciliation table)
  Sheet 2: Life of Contract Monthly Hours
  Sheet 3: All Time Entries (6,694 rows, life of contract)
  Sheet 4-6: Jan / Feb / Mar 2026 original xlsx data (as sent with monthly invoices)
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

HERE = Path(__file__).resolve().parent
INV_DIR = HERE / "monthly-invoice-emails"
OUT_PATH = HERE / "BWH-Hours-Accounting-Life-of-Contract.xlsx"

wb = openpyxl.Workbook()

# ===== Style helpers =====
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_col=None):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = min(len(s), 60)
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)


# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = "Summary"
ws1["A1"] = "Brandywine Homes — Contracted Hours Accounting"
ws1["A1"].font = Font(bold=True, size=16, color="1F4E78")
ws1.merge_cells("A1:F1")

ws1["A2"] = "Contract: IT Services Proposal (Contract ID 4924) | Client DirID: 6245"
ws1["A2"].font = Font(italic=True, color="505050")
ws1.merge_cells("A2:F2")
ws1["A3"] = f"Report generated: 2026-04-24 | Data source: Technijian Client Portal (direct SQL pull)"
ws1["A3"].font = Font(italic=True, color="505050")
ws1.merge_cells("A3:F3")

ws1["A5"] = "INVOICE-BY-INVOICE ROLL-FORWARD (as shown on monthly invoice footer)"
ws1["A5"].font = Font(bold=True, size=12)
ws1.merge_cells("A5:F5")

ws1["A6"] = "Formula: UnPaid (Aft) = UnPaid (Bef) + Actual(prev) - Billed"
ws1["A6"].font = Font(italic=True, color="505050")
ws1.merge_cells("A6:F6")

# Invoice 28148 footer table data (authoritative)
headers = ["Role", "UnPaid (Bef)", "Billed", "Actual(prev)", "UnPaid (Aft)", "Notes"]
inv_28148_rows = [
    ("India Tech: Normal",      572.20, 58.13, 59.74, 573.81, ""),
    ("India Tech: After Hours", 291.84, 42.82, 41.21, 290.23, ""),
    ("Systems_Architect.R",       0.00,  5.00,  0.00,   0.00, "5 hr allocation paid, 0 consumed"),
    ("USA Tech: Normal",        140.94, 15.26, 17.58, 143.26, ""),
]

ws1["A8"] = "Invoice #28148 (dated 4/1/2026, covering March 2026 support — latest monthly invoice)"
ws1["A8"].font = Font(bold=True, size=11, color="1F4E78")
ws1.merge_cells("A8:F8")

r = 9
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, len(headers))

r = 10
total_bef = total_bill = total_act = total_aft = 0.0
for row in inv_28148_rows:
    for i, v in enumerate(row, 1):
        c = ws1.cell(row=r, column=i, value=v)
        c.border = BORDER
        if i >= 2 and i <= 5:
            c.number_format = "0.00"
    total_bef += row[1]; total_bill += row[2]; total_act += row[3]; total_aft += row[4]
    r += 1

# Totals row
ws1.cell(row=r, column=1, value="TOTAL")
ws1.cell(row=r, column=2, value=total_bef)
ws1.cell(row=r, column=3, value=total_bill)
ws1.cell(row=r, column=4, value=total_act)
ws1.cell(row=r, column=5, value=total_aft)
for i in range(1, 7):
    c = ws1.cell(row=r, column=i)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER
    if i >= 2 and i <= 5:
        c.number_format = "0.00"

r += 2
ws1.cell(row=r, column=1, value="KEY NUMBERS")
ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
r += 1
summary_items = [
    ("Current unpaid balance (per Invoice #28148 footer, as of 3/31/2026)", f"{total_aft:.2f} hours"),
    ("Prior unpaid balance (per Invoice #28148 footer, as of 2/28/2026)", f"{total_bef:.2f} hours"),
    ("Contracted hours billed March 2026", f"{total_bill:.2f} hours"),
    ("Actual hours delivered February 2026 (under-contract)", f"{total_act:.2f} hours"),
    ("Life-of-contract hours delivered (2023-05 → 2026-04-24)", "4,050.10 hours"),
    ("Total weekly invoices sent (2023-05 → 2026-04-17)", "149 invoices"),
    ("Weekly invoices marked PAID in portal", "149 / 149 (100%)"),
]
for label, val in summary_items:
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=3, value=val)
    ws1.cell(row=r, column=1).font = Font(size=11)
    ws1.cell(row=r, column=3).font = Font(size=11, bold=True)
    r += 1

r += 1
ws1.cell(row=r, column=1, value="ABOUT THE FOOTER FORMAT CHANGE")
ws1.cell(row=r, column=1).font = Font(bold=True, size=12)
r += 1
notes = [
    "Beginning with Invoice #28148 (April 2026), we replaced the prior 'Support History (last 6 months)' footer",
    "with a clearer 4-column cumulative table: UnPaid (Bef) | Billed | Actual(prev) | UnPaid (Aft).",
    "",
    "The previous 6-month-window format only showed per-month deltas and did NOT clearly state the total",
    "carried-forward balance per the contract. The new format is intended to provide full transparency on the",
    "unpaid-hour balance per Section 2.05(a) of the Client Monthly Service Agreement.",
    "",
    "No data changed — only the presentation. The same balance has been accruing and has been reflected",
    "on the monthly invoices since the contract began; prior invoices simply did not display the running total.",
]
for note in notes:
    ws1.cell(row=r, column=1, value=note)
    ws1.cell(row=r, column=1).font = Font(size=10, italic=(note == ""))
    r += 1

# Set column widths
ws1.column_dimensions["A"].width = 60
ws1.column_dimensions["B"].width = 15
ws1.column_dimensions["C"].width = 15
ws1.column_dimensions["D"].width = 15
ws1.column_dimensions["E"].width = 15
ws1.column_dimensions["F"].width = 45
ws1.row_dimensions[1].height = 24


# ===== Sheet 2: Life of Contract Monthly =====
ws2 = wb.create_sheet("Life of Contract by Month")
ws2["A1"] = "Actual Hours Delivered by Month (2023-05 → 2026-04-24)"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws2.merge_cells("A1:F1")
ws2["A2"] = "Source: Client Portal stp_xml_TktEntry_List_Get (pulled 2026-04-24)"
ws2["A2"].font = Font(italic=True, color="505050")
ws2.merge_cells("A2:F2")

# Read hours-by-month.csv
with open(HERE / "hours-by-month.csv") as f:
    reader = csv.reader(f)
    data_rows = list(reader)

r = 4
for i, h in enumerate(data_rows[0], 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, len(data_rows[0]))

r = 5
for data_row in data_rows[1:]:
    for i, v in enumerate(data_row, 1):
        try:
            v = float(v) if i > 1 else v
        except ValueError:
            pass
        c = ws2.cell(row=r, column=i, value=v)
        c.border = BORDER
        if i > 1:
            c.number_format = "0.00"
    r += 1

# Totals
ws2.cell(row=r, column=1, value="TOTAL")
for i in range(2, 7):
    col_values = [ws2.cell(row=rr, column=i).value for rr in range(5, r)]
    total = sum(v for v in col_values if isinstance(v, (int, float)))
    c = ws2.cell(row=r, column=i, value=round(total, 2))
    c.number_format = "0.00"
for i in range(1, 7):
    c = ws2.cell(row=r, column=i)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER

ws2.column_dimensions["A"].width = 12
for col in ["B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 22


# ===== Sheet 3: All Time Entries (life of contract) =====
ws3 = wb.create_sheet("All Time Entries")
ws3["A1"] = "All BWH Time Entries — Life of Contract (2023-05 → 2026-04-24)"
ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws3.merge_cells("A1:J1")
ws3["A2"] = "Every row is one logged time entry from Technijian Client Portal, grouped by ticket"
ws3["A2"].font = Font(italic=True, color="505050")
ws3.merge_cells("A2:J2")

with open(HERE / "time-entries.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    all_te_rows = list(reader)

te_header = all_te_rows[0]
# Rename columns more friendly
friendly_cols = {
    "SourceMonth": "Source Month",
    "ConName": "Contract",
    "TimeEntryDate": "Date",
    "Title": "Ticket Title",
    "TimeDiff": "Time",
    "Resource": "Resource",
    "Requestor": "Requestor",
    "HourType": "Shift (N=Normal, AH=After Hours)",
    "BillRate": "Bill Rate",
    "PODDet": "POD Detail",
    "InvDescription": "Invoice Description",
    "InvDetID": "InvDetID",
    "StartDateTime": "Start",
    "EndDateTime": "End",
    "Office-POD": "POD",
    "RoleType": "Role",
    "WorkType": "Work Type",
    "AssignedName": "Assignee",
    "AH_Rate": "AH Rate",
    "NH_Rate": "NH Rate",
    "Hours": "Hours",
}
display_cols = ["Date", "Ticket Title", "Requestor", "POD", "Role",
                "Shift (N=Normal, AH=After Hours)", "Assignee", "Hours",
                "Invoice Description", "Source Month"]
col_index = {}
for i, h in enumerate(te_header):
    friendly = friendly_cols.get(h, h)
    col_index[friendly] = i

r = 4
for i, h in enumerate(display_cols, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, len(display_cols))

r = 5
for data_row in all_te_rows[1:]:
    for i, col_name in enumerate(display_cols, 1):
        src_idx = col_index.get(col_name)
        if src_idx is None:
            continue
        v = data_row[src_idx]
        if col_name == "Hours":
            try:
                v = abs(float(v))
            except (ValueError, TypeError):
                pass
        # Clean inv description
        if col_name == "Invoice Description":
            v = (v or "").replace("\r\n", " / ").replace("\n", " / ").strip(" /")
        c = ws3.cell(row=r, column=i, value=v)
        c.border = BORDER
        if col_name == "Hours":
            c.number_format = "0.00"
    r += 1

ws3.freeze_panes = "A5"
ws3.auto_filter.ref = f"A4:J{r-1}"
ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 60
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 22
ws3.column_dimensions["G"].width = 22
ws3.column_dimensions["H"].width = 10
ws3.column_dimensions["I"].width = 40
ws3.column_dimensions["J"].width = 12


# ===== Sheets 4-6: Monthly xlsx from invoice attachments =====
monthly_invoices = [
    ("27685", "Jan 2026 Tickets (Invoice 27685)", "January 2026"),
    ("27948", "Feb 2026 Tickets (Invoice 27948)", "February 2026"),
    ("28148", "Mar 2026 Tickets (Invoice 28148)", "March 2026"),
]

for inv_no, sheet_title, month_label in monthly_invoices:
    src_path = INV_DIR / f"BWH  Monthly Invoice {inv_no} for your review__Monthly_Time_Entries_for_{inv_no}.xlsx"
    src_wb = openpyxl.load_workbook(src_path, data_only=True)
    src_ws = src_wb.active

    ws = wb.create_sheet(sheet_title)
    ws["A1"] = f"{month_label} — Time Entries (as attached to Monthly Invoice #{inv_no})"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:R1")
    ws["A2"] = f"Original file: Monthly_Time_Entries_for_{inv_no}.xlsx"
    ws["A2"].font = Font(italic=True, color="505050")
    ws.merge_cells("A2:R2")

    src_rows = list(src_ws.iter_rows(values_only=True))
    if not src_rows:
        continue
    hdr = src_rows[0]
    r = 4
    for i, h in enumerate(hdr, 1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, len(hdr))

    r = 5
    for row in src_rows[1:]:
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORDER
        r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(hdr))}{r-1}"
    # Column widths
    for i in range(1, len(hdr) + 1):
        col = get_column_letter(i)
        hname = (hdr[i-1] or "").lower()
        if "title" in hname or "note" in hname:
            ws.column_dimensions[col].width = 50
        elif "date" in hname or "start" in hname or "end" in hname:
            ws.column_dimensions[col].width = 18
        else:
            ws.column_dimensions[col].width = 15


wb.save(OUT_PATH)
print(f"Wrote: {OUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
