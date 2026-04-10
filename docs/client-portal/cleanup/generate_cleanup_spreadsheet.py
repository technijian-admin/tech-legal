import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("c:/vscode/tech-legal/active-clients-signers-directory.json") as f:
    data = json.load(f)

wb = Workbook()

# Color palette (Technijian brand)
BLUE = "006DB6"
ORANGE = "F67D4B"
DARK = "1A1A2E"
TEAL = "1EAAC8"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_BG = "FFCCCC"
ORANGE_BG = "FFE0CC"
YELLOW_BG = "FFFFCC"
GREEN_BG = "CCFFCC"

header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
title_font = Font(name="Calibri", size=14, bold=True, color=DARK)
subtitle_font = Font(name="Calibri", size=10, color="666666")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
issue_font = Font(name="Calibri", size=10, color="CC0000")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

red_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
orange_fill = PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def detect_issues(entry):
    """Return list of (severity, issue_text, recommendation) tuples."""
    issues = []
    signers = entry.get("authorizedSigners", [])
    addr = entry.get("address", {})
    name = entry.get("clientName") or ""

    # Critical
    if not entry.get("hasAuthorizedSigner"):
        issues.append(("Critical", "No authorized signers (C1 or C2)", "Add at least one C1 Admin user"))
    if name.lower() in ("none", "") or "test" in name.lower():
        issues.append(("Critical", f'Possible test/dummy client: "{name}"', "Remove or archive if not real"))

    # Warning
    if entry.get("hasAuthorizedSigner") and entry.get("signerCount", 0) == 0:
        issues.append(("Warning", "No C1 Admin signer (C2 Manager only)", "Promote a C2 to C1 or add C1"))
    if not entry.get("directoryActive"):
        issues.append(("Warning", "Directory INACTIVE but has active contracts", "Set IsActive=true or end-date contracts"))
    if signers:
        pname = signers[0].get("fullName", "")
        if pname and pname.strip().lower() in ("accounting/billing", "billing", "admin", "it", "office", "reception"):
            issues.append(("Warning", f'Primary signer is generic: "{pname}"', "Replace with named individual"))

    # Info
    if not addr.get("line1") and not addr.get("city"):
        issues.append(("Info", "No address on file", "Update client address"))
    if not entry.get("phone"):
        issues.append(("Info", "No phone on file", "Update client phone"))
    if not entry.get("domain"):
        issues.append(("Info", "No domain on file", "Add client email domain"))

    return issues


def max_severity(issues):
    if any(s == "Critical" for s, _, _ in issues):
        return 3
    if any(s == "Warning" for s, _, _ in issues):
        return 2
    if any(s == "Info" for s, _, _ in issues):
        return 1
    return 0


def severity_fill(sev):
    if sev == 3 or sev == "Critical":
        return red_fill
    if sev == 2 or sev == "Warning":
        return orange_fill
    if sev == 1 or sev == "Info":
        return yellow_fill
    return None


# ═══════════════════════════════════════════
# Sheet 1: Active Clients Directory
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "Active Clients Directory"
ws.sheet_properties.tabColor = BLUE

ws.merge_cells("A1:N1")
ws["A1"] = "Active Clients & Authorized Signers Directory"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:N2")
ws["A2"] = (
    f"Generated {data['generatedAt']} | "
    "Active = contract started, no end date, marked active | "
    "Signers = C1 (Admin) + C2 (Manager)"
)
ws["A2"].font = subtitle_font
ws.row_dimensions[2].height = 18

# Legend row
ws.merge_cells("A3:B3")
ws["A3"] = "Legend:"
ws["A3"].font = bold_font
for ci, (label, fill) in enumerate(
    [("Critical", red_fill), ("Warning", orange_fill), ("Info", yellow_fill), ("Clean", green_fill)], 3
):
    cell = ws.cell(row=3, column=ci, value=label)
    cell.fill = fill
    cell.font = normal_font
    cell.border = thin_border

headers = [
    ("Client Code", 13),
    ("Client Name", 32),
    ("Dir Active", 11),
    ("# Contracts", 12),
    ("# C1 Signers", 12),
    ("# C1+C2", 10),
    ("Has Signer?", 11),
    ("Primary Signer", 24),
    ("Signer Email", 30),
    ("Signer Title", 20),
    ("Phone", 18),
    ("City, State", 22),
    ("Domain", 22),
    ("Issues", 50),
]

for col_idx, (header, width) in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[5].height = 28
ws.auto_filter.ref = "A5:N5"
ws.freeze_panes = "A6"

row = 6
for entry in sorted(data["directory"], key=lambda e: (e.get("clientName") or "")):
    issues = detect_issues(entry)
    sev = max_severity(issues)
    signers = entry.get("authorizedSigners", [])
    primary = signers[0] if signers else {}
    addr = entry.get("address", {})
    city_state = ", ".join(filter(None, [addr.get("city"), addr.get("state")])) or ""

    values = [
        entry.get("clientCode"),
        entry.get("clientName"),
        "Yes" if entry.get("directoryActive") else "No",
        len(entry.get("activeContracts", [])),
        entry.get("signerCount", 0),
        len(signers),
        "Yes" if entry.get("hasAuthorizedSigner") else "NO",
        primary.get("fullName", ""),
        primary.get("email", ""),
        primary.get("title", "") or "",
        entry.get("phone") or "",
        city_state,
        entry.get("domain") or "",
        "; ".join(f"[{s}] {t}" for s, t, _ in issues) if issues else "Clean",
    ]

    row_fill = severity_fill(sev)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = issue_font if (col_idx == 14 and issues) else normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 14))
        if row_fill:
            cell.fill = row_fill
        elif not issues and row % 2 == 0:
            cell.fill = alt_fill
        elif not issues:
            cell.fill = green_fill

    ws.row_dimensions[row].height = 20 if not issues else 32
    row += 1


# ═══════════════════════════════════════════
# Sheet 2: Issues Summary (prioritized)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Issues to Fix")
ws2.sheet_properties.tabColor = ORANGE

ws2.merge_cells("A1:F1")
ws2["A1"] = "Client Portal Data Quality Issues"
ws2["A1"].font = title_font
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:F2")
ws2["A2"] = "Sorted by severity: Critical > Warning > Info. Fix top rows first."
ws2["A2"].font = subtitle_font

issue_headers = [
    ("Client Code", 14),
    ("Client Name", 30),
    ("Severity", 12),
    ("Issue", 50),
    ("# Contracts", 12),
    ("Recommendation", 45),
]
for col_idx, (h, w) in enumerate(issue_headers, 1):
    cell = ws2.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

ws2.auto_filter.ref = "A4:F4"
ws2.freeze_panes = "A5"

# Collect all issue rows
all_issue_rows = []
for entry in data["directory"]:
    issues = detect_issues(entry)
    for sev, issue, rec in issues:
        all_issue_rows.append((
            sev,
            entry.get("clientCode"),
            entry.get("clientName"),
            issue,
            len(entry.get("activeContracts", [])),
            rec,
        ))

sev_order = {"Critical": 0, "Warning": 1, "Info": 2}
all_issue_rows.sort(key=lambda r: (sev_order.get(r[0], 9), r[2] or ""))

row2 = 5
for sev, code, name, issue, ncontracts, rec in all_issue_rows:
    fill = severity_fill(sev)
    vals = [code, name, sev, issue, ncontracts, rec]
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row2, column=col_idx, value=val)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (4, 6)))
    ws2.row_dimensions[row2].height = 30
    row2 += 1


# ═══════════════════════════════════════════
# Sheet 3: All Signers
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("All Signers")
ws3.sheet_properties.tabColor = TEAL

ws3.merge_cells("A1:H1")
ws3["A1"] = "All Authorized Signers by Client"
ws3["A1"].font = title_font
ws3.row_dimensions[1].height = 30

signer_headers = [
    ("Client Code", 12),
    ("Client Name", 28),
    ("Signer Name", 25),
    ("Email", 30),
    ("Title", 20),
    ("Phone", 18),
    ("Role", 8),
    ("Role Description", 16),
]
for col_idx, (h, w) in enumerate(signer_headers, 1):
    cell = ws3.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws3.column_dimensions[get_column_letter(col_idx)].width = w

ws3.auto_filter.ref = "A3:H3"
ws3.freeze_panes = "A4"

row3 = 4
for entry in sorted(data["directory"], key=lambda e: (e.get("clientName") or "")):
    for s in entry.get("authorizedSigners", []):
        vals = [
            entry.get("clientCode"),
            entry.get("clientName"),
            s.get("fullName"),
            s.get("email"),
            s.get("title") or "",
            s.get("phone") or "",
            s.get("role"),
            s.get("roleDescription"),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row3, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row3 % 2 == 0:
                cell.fill = alt_fill
        row3 += 1


# ═══════════════════════════════════════════
# Sheet 4: Dashboard
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard")
ws4.sheet_properties.tabColor = BLUE
ws4.column_dimensions["A"].width = 36
ws4.column_dimensions["B"].width = 12
ws4.column_dimensions["C"].width = 10
ws4.column_dimensions["D"].width = 18

ws4.merge_cells("A1:D1")
ws4["A1"] = "Client Portal Cleanup Dashboard"
ws4["A1"].font = title_font
ws4.row_dimensions[1].height = 30

summary = data["summary"]
total = summary["totalActiveClients"]

stat_headers = ["Metric", "Value", "%", "Status"]
for col_idx, h in enumerate(stat_headers, 1):
    cell = ws4.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

stats = [
    ("Total Active Clients", summary["totalActiveClients"], "", ""),
    ("Total Active Contracts", summary["totalActiveContracts"], "", ""),
    ("Total C1 Signers (Admin)", summary["totalC1Signers"], "", ""),
    ("Total C1+C2 Signers", summary["totalC1C2Signers"], "", ""),
    ("", "", "", ""),
    ("Clients WITH Signers", summary["clientsWithSigners"], f"{summary['clientsWithSigners']/total*100:.0f}%", "Good"),
    ("Clients WITHOUT Signers", summary["clientsWithoutSigners"], f"{summary['clientsWithoutSigners']/total*100:.0f}%", "Action needed"),
    ("Clients INACTIVE in Directory", summary["clientsInactiveInDirectory"], f"{summary['clientsInactiveInDirectory']/total*100:.0f}%", "Review needed"),
]

for i, (metric, val, pct, status) in enumerate(stats, 4):
    ws4.cell(row=i, column=1, value=metric).font = bold_font if metric else normal_font
    ws4.cell(row=i, column=2, value=val).font = normal_font
    ws4.cell(row=i, column=3, value=pct).font = normal_font
    ws4.cell(row=i, column=4, value=status).font = normal_font
    for c in range(1, 5):
        ws4.cell(row=i, column=c).border = thin_border
        if "Action" in status:
            ws4.cell(row=i, column=c).fill = red_fill
        elif "Review" in status:
            ws4.cell(row=i, column=c).fill = orange_fill
        elif "Good" in status:
            ws4.cell(row=i, column=c).fill = green_fill

# Issue breakdown
ws4.cell(row=14, column=1, value="Issue Breakdown").font = bold_font
for col_idx, h in enumerate(["Issue Type", "Count"], 1):
    cell = ws4.cell(row=15, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

issue_type_counts = {}
for sev, code, name, issue, nc, rec in all_issue_rows:
    key = issue
    if key not in issue_type_counts:
        issue_type_counts[key] = {"count": 0, "severity": sev}
    issue_type_counts[key]["count"] += 1

r = 16
for issue_type, info in sorted(issue_type_counts.items(), key=lambda x: (-sev_order.get(x[1]["severity"], 9), -x[1]["count"])):
    ws4.cell(row=r, column=1, value=issue_type).font = normal_font
    ws4.cell(row=r, column=2, value=info["count"]).font = normal_font
    fill = severity_fill(info["severity"])
    for c in range(1, 3):
        ws4.cell(row=r, column=c).border = thin_border
        if fill:
            ws4.cell(row=r, column=c).fill = fill
    ws4.column_dimensions["A"].width = 52
    r += 1


# Save
wb.save("c:/vscode/tech-legal/active-clients-signers-cleanup.xlsx")
print("Saved: active-clients-signers-cleanup.xlsx")
print(f"  Sheet 1: {row - 6} clients")
print(f"  Sheet 2: {len(all_issue_rows)} issue rows")
print(f"  Sheet 3: {row3 - 4} signer rows")
print(f"  Sheet 4: Dashboard with {len(issue_type_counts)} issue types")
